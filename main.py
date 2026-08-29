from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from time import sleep
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

movie_dict = {
  "name":[],
  "year":[],
  "duration":[],
  "stars":[],
  "votes":[],
  "metascore":[],
  "description":[]
}

options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
options.add_experimental_option('detach', True)


url = "https://www.imdb.com/"
driver = webdriver.Chrome(options=options)
actions = ActionChains(driver)
driver.implicitly_wait(1)
driver.get(url)
print("***Please complete the manual Human verification process***")
input("Once IMDB has loaded, press enter to continue!")

try:
  WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "suggestion-search-button")))
  search_btn = driver.find_element(By.ID, "suggestion-search-button")
  search_btn.click()
except Exception as e:
  print(f"Error: {e}")

try:
  WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//a[@data-testid='advanced-search-chip-tt']")))
  advanced_search= driver.find_element(By.XPATH, "//a[@data-testid='advanced-search-chip-tt']")
  advanced_search.click()
except Exception as e:
  print(f"Error: {e}")

try:
  WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//button[@data-testid='test-chip-id-movie']")))
  movie_btn= driver.find_element(By.XPATH, "//button[@data-testid='test-chip-id-movie']")
  movie_btn.click()
except Exception as e:
  print(f"Error: {e}")

try:
  genre_div=driver.find_element(By.XPATH, ("//div[text()='Genre']"))
  ActionChains(driver).move_to_element(genre_div).click().perform()

  sleep(10)
  ActionChains(driver).move_to_element(genre_div).click().perform()

  WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//button[@data-testid='test-chip-id-Comedy']")))
  comedy_btn = driver.find_element(By.XPATH, "//button[@data-testid='test-chip-id-Comedy']")
  ActionChains(driver).move_to_element(comedy_btn).click().perform()
except Exception as e:
  print(f"Error: {e}")

try:
  awards_div = driver.find_element(By.XPATH, "//div[text()='Awards & recognition']")
  sleep(5)
  ActionChains(driver).move_to_element(awards_div).click().perform()
  WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//button[@data-testid='test-chip-id-oscar-nominated']")))
  oscar_btn= driver.find_element(By.XPATH, "//button[@data-testid='test-chip-id-oscar-nominated']")
  ActionChains(driver).move_to_element(oscar_btn).click().perform()
except Exception as e:
  print(f"Error: {e}")

try:
  results_btn = driver.find_element(By.XPATH, "//button[@data-testid='adv-search-get-results']")
  results_btn.click()
except Exception as e:
  print(f"Error: {e}")

while True:
  sleep(2)
  more_btns = driver.find_elements(By.XPATH, "//button[.//span[@class='ipc-see-more__text']]")
  if len(more_btns):
    more_btn = more_btns[0]
    ActionChains(driver).move_to_element(more_btn).click().perform()
  else:
    break

movies = driver.find_elements(By.CLASS_NAME, "ipc-metadata-list-summary-item")
for i, movie in enumerate(movies, start=1):
  print(f"Processing movie {i}/{len(movies)}", flush=True)
  raw_name = movie.find_element(By.XPATH, ".//h4[@class='ipc-title__text']").text
  name = " ".join(raw_name.split(" ")[1:])
  movie_dict["name"].append(name)
  year_n_duration=movie.find_elements(By.CSS_SELECTOR, "li.ipc-inline-list__item")
  year = year_n_duration[0].text
  duration = year_n_duration[1].text
  movie_dict["year"].append(year)
  movie_dict["duration"].append(duration)
  star_rating = movie.find_element(By.CSS_SELECTOR, "span.ipc-rating-star--rating").text
  movie_dict["stars"].append(star_rating)
  votes_raw = movie.find_element(By.XPATH, ".//span[@class='ipc-rating-star--voteCount']").text
  votes = votes_raw.strip().strip("()")
  movie_dict["votes"].append(votes)
  try:
        metascore = movie.find_element(By.XPATH, ".//span[@class='sc-9fe7b0ef-0 hDuMnh metacritic-score-box']").text
        movie_dict['metascore'].append(metascore)
  except:
        metascore = 'No info'
        movie_dict['metascore'].append(metascore)
  description = movie.find_element(By.CLASS_NAME, "ipc-html-content-inner-div").text
  movie_dict["description"].append(description)

df = pd.DataFrame(movie_dict)
df.insert(0, "No", range(1, len(df)+1))
df.columns = [column.capitalize() for column in df.columns]
df.to_excel("movies.xlsx", index=False)

wb = load_workbook("movies.xlsx")
ws =wb.active

for cell in ws[1]:
  cell.font = Font(bold=True)
wb.save("movies.xlsx")






